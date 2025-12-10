"""
Integration Tests for Excel Generation

Test complete Excel generation workflow.
"""

import pytest
import sys
from pathlib import Path

# Add fastapi directory to path BEFORE importing captable
test_dir = Path(__file__).resolve().parent.parent
fastapi_dir = test_dir.parent / "fastapi"
if str(fastapi_dir) not in sys.path:
    sys.path.insert(0, str(fastapi_dir))

from captable.generator import CapTableGenerator


class TestExcelGeneration:
    """Test Excel generation."""

    def test_generate_excel_complete_cap_table(self, temp_dir, sample_cap_table_data):
        """Test generating Excel from complete cap table."""
        output_path = temp_dir / "test_output.xlsx"

        generator = CapTableGenerator(json_data=sample_cap_table_data)
        assert generator.validate()

        result_path = generator.generate_excel(str(output_path))
        assert Path(result_path).exists()
        assert result_path == str(output_path)

    def test_generate_excel_all_calculation_types(self, temp_dir):
        """Test generating Excel with all calculation types."""
        from tests.fixtures.sample_cap_tables import (
            FIXED_SHARES_ROUND,
            TARGET_PERCENTAGE_ROUND,
            VALUATION_BASED_ROUND,
            CONVERTIBLE_ROUND,
            SAFE_ROUND,
        )

        data = {
            "schema_version": "2.0",
            "holders": [
                {"name": "Founder 1", "group": "Founders"},
                {"name": "Investor A", "group": "Investors"},
            ],
            "rounds": [
                FIXED_SHARES_ROUND,
                TARGET_PERCENTAGE_ROUND,
                VALUATION_BASED_ROUND,
                CONVERTIBLE_ROUND,
                SAFE_ROUND,
            ],
        }

        output_path = temp_dir / "test_all_types.xlsx"
        generator = CapTableGenerator(json_data=data)

        if generator.validate():
            result_path = generator.generate_excel(str(output_path))
            assert Path(result_path).exists()

    def test_generate_excel_with_anti_dilution(self, temp_dir):
        """Test generating Excel with anti-dilution protection."""
        data = {
            "schema_version": "2.0",
            "holders": [
                {"name": "Founder 1", "group": "Founders"},
                {"name": "Investor A", "group": "Investors"},
                {"name": "Investor B", "group": "Investors"},
                {"name": "Investor C", "group": "Investors"},
            ],
            "rounds": [
                {
                    "name": "Seed Round",
                    "calculation_type": "valuation_based",
                    "round_date": "2024-01-01",
                    "valuation": 1000000,
                    "valuation_basis": "pre_money",
                    "instruments": [
                        {
                            "holder_name": "Investor A",
                            "class_name": "Preferred",
                            "investment_amount": 200000,
                            "dilution_method": "full_ratchet",
                        },
                        {
                            "holder_name": "Investor B",
                            "class_name": "Preferred",
                            "investment_amount": 300000,
                            "dilution_method": "narrow_based_weighted_average",
                        },
                    ],
                },
                {
                    "name": "Series A",
                    "calculation_type": "valuation_based",
                    "round_date": "2024-06-01",
                    "valuation": 800000,  # Lower valuation triggers anti-dilution
                    "valuation_basis": "pre_money",
                    "instruments": [
                        {
                            "holder_name": "Investor C",
                            "class_name": "Preferred",
                            "investment_amount": 100000,
                        },
                    ],
                },
            ],
        }

        output_path = temp_dir / "test_anti_dilution.xlsx"
        generator = CapTableGenerator(json_data=data)

        assert generator.validate()

        result_path = generator.generate_excel(str(output_path))
        assert Path(result_path).exists()

        # Verify the Excel file contains the Anti-Dilution Allocations sheet
        import openpyxl
        workbook = openpyxl.load_workbook(result_path)
        sheet_names = workbook.sheetnames
        assert "Anti-Dilution Allocations" in sheet_names

        # Verify the sheet has data
        anti_dilution_sheet = workbook["Anti-Dilution Allocations"]
        # Check that the sheet has content (more than just headers)
        assert anti_dilution_sheet.max_row > 3  # Title + headers + at least one data row
