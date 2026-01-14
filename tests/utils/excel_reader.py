"""
Excel Reader

Utilities for reading and parsing generated Excel files.
"""

from pathlib import Path
from typing import Dict, Any, List, Optional
import openpyxl  # type: ignore


class ExcelReader:
    """Read and parse Excel files for testing."""
    
    def __init__(self, file_path: str | Path):
        """
        Initialize Excel reader.
        
        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[openpyxl.Workbook] = None
    
    def load(self) -> None:
        """Load Excel workbook."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        if not self.workbook:
            self.load()
        return self.workbook.sheetnames  # type: ignore
    
    def get_sheet(self, sheet_name: str) -> Optional[Any]:
        """Get worksheet by name."""
        if not self.workbook:
            self.load()
        return self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else None  # type: ignore
    
    def read_cell(self, sheet_name: str, cell_address: str) -> Any:
        """
        Read cell value.
        
        Args:
            sheet_name: Name of sheet
            cell_address: Cell address (e.g., "A1")
            
        Returns:
            Cell value
        """
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return sheet[cell_address].value
    
    def read_range(self, sheet_name: str, range_address: str) -> List[List[Any]]:
        """
        Read range of cells.
        
        Args:
            sheet_name: Name of sheet
            range_address: Range address (e.g., "A1:B10")
            
        Returns:
            List of rows, each row is a list of cell values
        """
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            raise ValueError(f"Sheet not found: {sheet_name}")
        
        cells = sheet[range_address]
        if isinstance(cells, tuple):
            # Single row
            return [[cell.value for cell in cells]]
        else:
            # Multiple rows
            return [[cell.value for cell in row] for row in cells]
    
    def get_named_ranges(self) -> Dict[str, str]:
        """Get all named ranges in workbook."""
        if not self.workbook:
            self.load()
        
        named_ranges = {}
        for name, definition in self.workbook.defined_names.items():  # type: ignore
            named_ranges[name] = str(definition)
        
        return named_ranges
    
    def close(self) -> None:
        """Close workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None





