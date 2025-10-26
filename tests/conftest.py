"""
Pytest Configuration and Fixtures
Provides reusable test fixtures for cap table generation tests.
"""

import sys
from pathlib import Path

# Add project root to Python path so we can import from src
project_root = Path(__file__).parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import pytest
import json
import tempfile
import os
from typing import Dict, Any
import openpyxl
from openpyxl.workbook import Workbook


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test outputs."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def simple_holders():
    """Fixture: Simple holder data."""
    return [
        {"name": "Alice", "type": "founder", "email": "alice@test.com"},
        {"name": "Bob", "type": "investor", "email": "bob@test.com"}
    ]


@pytest.fixture
def simple_classes():
    """Fixture: Simple class data."""
    return [
        {"name": "Common", "type": "common"}
    ]


@pytest.fixture
def simple_terms():
    """Fixture: Simple terms data."""
    return [
        {
            "name": "Standard Terms",
            "class_name": "Preferred",
            "liquidation_multiple": 1.0,
            "participation_type": "non_participating",
            "seniority_rank": 1
        }
    ]


@pytest.fixture
def simple_instruments():
    """Fixture: Simple instrument data (legacy format for testing)."""
    return [
        {
            "holder_name": "Alice",
            "class_name": "Common",
            "initial_quantity": 1000000,
            "acquisition_price": 0.001,
            "acquisition_date": "2023-01-01"
        },
        {
            "holder_name": "Bob",
            "class_name": "Common",
            "initial_quantity": 500000,
            "acquisition_price": 0.001,
            "acquisition_date": "2023-01-01"
        }
    ]


@pytest.fixture
def simple_rounds():
    """Fixture: Simple round data (legacy format for testing)."""
    return [
        {
            "name": "Seed",
            "round_date": "2023-06-01",
            "investment_amount": 500000,
            "pre_money_valuation": 1500000,
            "post_money_valuation": 2000000,
            "price_per_share": 1.0
        }
    ]


@pytest.fixture
def simple_rounds_v2():
    """Fixture: Simple rounds data (v2.0 format with nested instruments)."""
    return [
        {
            "name": "Incorporation",
            "round_date": "2023-01-01",
            "calculation_type": "fixed_shares",
            "instruments": [
                {
                    "holder_name": "Alice",
                    "class_name": "Common",
                    "initial_quantity": 1000000,
                    "acquisition_date": "2023-01-01"
                },
                {
                    "holder_name": "Bob",
                    "class_name": "Common",
                    "initial_quantity": 500000,
                    "acquisition_date": "2023-01-01"
                }
            ]
        },
        {
            "name": "Seed",
            "round_date": "2023-06-01",
            "calculation_type": "valuation_based",
            "pre_money_valuation": 1500000,
            "post_money_valuation": 2000000,
            "price_per_share": 1.0,
            "instruments": [
                {
                    "holder_name": "Seed Investor",
                    "class_name": "Preferred",
                    "investment_amount": 500000,
                    "accrued_interest": 0
                }
            ]
        }
    ]


@pytest.fixture
def minimal_cap_table(simple_holders, simple_classes, simple_instruments):
    """Fixture: Minimal valid cap table (legacy format for testing)."""
    return {
        "schema_version": "1.0",
        "company": {
            "name": "Test Company",
            "current_date": "2024-01-01",
            "current_pps": 1.0
        },
        "holders": simple_holders,
        "classes": simple_classes,
        "instruments": simple_instruments
    }


@pytest.fixture
def minimal_cap_table_v2():
    """Fixture: Minimal valid cap table (v2.0 format)."""
    return {
        "schema_version": "2.0",
        "company": {
            "name": "Test Company",
            "current_date": "2024-01-01"
        },
        "rounds": [
            {
                "name": "Incorporation",
                "round_date": "2023-01-01",
                "calculation_type": "fixed_shares",
                "instruments": [
                    {
                        "holder_name": "Alice",
                        "class_name": "Common",
                        "initial_quantity": 1000000,
                        "acquisition_date": "2023-01-01"
                    }
                ]
            }
        ]
    }


@pytest.fixture
def full_cap_table(simple_holders, simple_classes, simple_terms, simple_instruments, simple_rounds):
    """Fixture: Full cap table with all sections (legacy format for testing)."""
    return {
        "schema_version": "1.0",
        "company": {
            "name": "Test Company",
            "incorporation_date": "2023-01-01",
            "current_date": "2024-01-01",
            "current_pps": 2.0
        },
        "holders": simple_holders,
        "classes": simple_classes,
        "terms": simple_terms,
        "instruments": simple_instruments,
        "rounds": simple_rounds,
        "waterfall_scenarios": [
            {
                "name": "Exit at $10M",
                "exit_value": 10000000
            }
        ]
    }


@pytest.fixture
def full_cap_table_v2(simple_rounds_v2):
    """Fixture: Full cap table (v2.0 format with all calculation types)."""
    return {
        "schema_version": "2.0",
        "company": {
            "name": "Test Company",
            "incorporation_date": "2023-01-01",
            "current_date": "2024-01-01"
        },
        "rounds": simple_rounds_v2
    }


@pytest.fixture
def complex_cap_table():
    """Fixture: Complex cap table with multiple rounds, options, vesting."""
    return {
        "schema_version": "1.0",
        "company": {
            "name": "Complex Corp",
            "incorporation_date": "2022-01-01",
            "current_date": "2024-01-01",
            "current_pps": 5.0
        },
        "holders": [
            {"name": "Founder1", "type": "founder", "email": "f1@test.com"},
            {"name": "Founder2", "type": "founder", "email": "f2@test.com"},
            {"name": "Employee1", "type": "employee", "email": "e1@test.com"},
            {"name": "Investor1", "type": "investor", "email": "i1@test.com"},
            {"name": "Investor2", "type": "investor", "email": "i2@test.com"},
            {"name": "Option Pool", "type": "option_pool"}
        ],
        "classes": [
            {"name": "Common", "type": "common"},
            {"name": "Options", "type": "option"},
            {"name": "Series A", "type": "preferred", "terms_name": "Series A Terms"},
            {"name": "Series B", "type": "preferred", "terms_name": "Series B Terms"}
        ],
        "terms": [
            {
                "name": "Series A Terms",
                "class_name": "Series A",
                "liquidation_multiple": 1.0,
                "participation_type": "non_participating",
                "seniority_rank": 2
            },
            {
                "name": "Series B Terms",
                "class_name": "Series B",
                "liquidation_multiple": 1.5,
                "participation_type": "participating",
                "seniority_rank": 1
            }
        ],
        "instruments": [
            {"holder_name": "Founder1", "class_name": "Common", "initial_quantity": 5000000, "acquisition_price": 0.001, "acquisition_date": "2022-01-01"},
            {"holder_name": "Founder2", "class_name": "Common", "initial_quantity": 5000000, "acquisition_price": 0.001, "acquisition_date": "2022-01-01"},
            {"holder_name": "Option Pool", "class_name": "Options", "initial_quantity": 2000000, "strike_price": 0.1, "acquisition_date": "2022-01-01"},
            {
                "holder_name": "Employee1",
                "class_name": "Options",
                "initial_quantity": 100000,
                "strike_price": 1.0,
                "acquisition_date": "2022-06-01",
                "vesting_terms": {
                    "grant_date": "2022-06-01",
                    "cliff_days": 365,
                    "vesting_period_days": 1460
                }
            },
            {"holder_name": "Investor1", "class_name": "Series A", "round_name": "Series A", "initial_quantity": 2000000, "acquisition_price": 2.0, "acquisition_date": "2023-01-01"},
            {"holder_name": "Investor2", "class_name": "Series B", "round_name": "Series B", "initial_quantity": 1000000, "acquisition_price": 4.0, "acquisition_date": "2023-07-01"}
        ],
        "rounds": [
            {
                "name": "Series A",
                "round_date": "2023-01-01",
                "investment_amount": 4000000,
                "pre_money_valuation": 20000000,
                "post_money_valuation": 24000000,
                "price_per_share": 2.0
            },
            {
                "name": "Series B",
                "round_date": "2023-07-01",
                "investment_amount": 4000000,
                "pre_money_valuation": 44000000,
                "post_money_valuation": 48000000,
                "price_per_share": 4.0
            }
        ],
        "waterfall_scenarios": [
            {"name": "Exit at $50M", "exit_value": 50000000}
        ]
    }


class ExcelHelper:
    """Helper class for Excel file assertions and inspections."""
    
    @staticmethod
    def load_workbook(file_path: Path) -> Workbook:
        """Load an Excel workbook."""
        return openpyxl.load_workbook(file_path, data_only=False)
    
    @staticmethod
    def get_sheet_names(file_path: Path) -> list:
        """Get list of sheet names in workbook."""
        wb = ExcelHelper.load_workbook(file_path)
        return wb.sheetnames
    
    @staticmethod
    def sheet_exists(file_path: Path, sheet_name: str) -> bool:
        """Check if a sheet exists in workbook."""
        return sheet_name in ExcelHelper.get_sheet_names(file_path)
    
    @staticmethod
    def get_cell_value(file_path: Path, sheet_name: str, cell: str):
        """Get value of a specific cell."""
        wb = ExcelHelper.load_workbook(file_path)
        ws = wb[sheet_name]
        return ws[cell].value
    
    @staticmethod
    def get_cell_formula(file_path: Path, sheet_name: str, cell: str):
        """Get formula of a specific cell."""
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        cell_obj = ws[cell]
        # Return the formula if it exists, otherwise the value
        if cell_obj.data_type == 'f':
            # Handle ArrayFormula objects
            formula = cell_obj.value
            if hasattr(formula, 'text'):
                return formula.text
            elif isinstance(formula, str):
                return formula
            else:
                return str(formula)
        return None
    
    @staticmethod
    def is_formula(file_path: Path, sheet_name: str, cell: str) -> bool:
        """Check if a cell contains a formula."""
        formula = ExcelHelper.get_cell_formula(file_path, sheet_name, cell)
        return formula is not None and formula.startswith('=')
    
    @staticmethod
    def get_table_names(file_path: Path, sheet_name: str) -> list:
        """Get list of table names in a sheet."""
        wb = ExcelHelper.load_workbook(file_path)
        ws = wb[sheet_name]
        return [table.name for table in ws.tables.values()]
    
    @staticmethod
    def table_exists(file_path: Path, table_name: str) -> bool:
        """Check if a table exists in any sheet."""
        wb = ExcelHelper.load_workbook(file_path)
        for ws in wb.worksheets:
            if table_name in [table.name for table in ws.tables.values()]:
                return True
        return False
    
    @staticmethod
    def get_named_ranges(file_path: Path) -> dict:
        """Get all named ranges in workbook."""
        wb = ExcelHelper.load_workbook(file_path)
        return {name: str(wb.defined_names[name].attr_text) for name in wb.defined_names}
    
    @staticmethod
    def named_range_exists(file_path: Path, name: str) -> bool:
        """Check if a named range exists."""
        return name in ExcelHelper.get_named_ranges(file_path)
    
    @staticmethod
    def get_data_validations(file_path: Path, sheet_name: str) -> list:
        """Get data validations in a sheet."""
        wb = ExcelHelper.load_workbook(file_path)
        ws = wb[sheet_name]
        return list(ws.data_validations.dataValidation)
    
    @staticmethod
    def has_data_validation(file_path: Path, sheet_name: str, cell: str) -> bool:
        """Check if a cell has data validation."""
        wb = ExcelHelper.load_workbook(file_path)
        ws = wb[sheet_name]
        cell_obj = ws[cell]
        
        for dv in ws.data_validations.dataValidation:
            if cell_obj.coordinate in dv.cells:
                return True
        return False


@pytest.fixture
def excel_helper():
    """Fixture: Excel helper utilities."""
    return ExcelHelper()


@pytest.fixture
def sample_json_file(temp_dir, full_cap_table):
    """Fixture: Create a sample JSON file."""
    json_path = temp_dir / "sample.json"
    with open(json_path, 'w') as f:
        json.dump(full_cap_table, f, indent=2)
    return json_path


def assert_excel_structure(file_path: Path, expected_sheets: list):
    """Assert that Excel file has expected structure."""
    assert file_path.exists(), f"Excel file not found: {file_path}"
    
    helper = ExcelHelper()
    actual_sheets = helper.get_sheet_names(file_path)
    
    for sheet in expected_sheets:
        assert sheet in actual_sheets, f"Expected sheet '{sheet}' not found. Found: {actual_sheets}"


def assert_formula_in_cell(file_path: Path, sheet_name: str, cell: str, expected_function: str = None):
    """Assert that a cell contains a formula."""
    helper = ExcelHelper()
    assert helper.is_formula(file_path, sheet_name, cell), \
        f"Cell {sheet_name}!{cell} does not contain a formula"
    
    if expected_function:
        formula = helper.get_cell_formula(file_path, sheet_name, cell)
        assert expected_function in formula, \
            f"Formula in {sheet_name}!{cell} does not contain {expected_function}. Found: {formula}"


def assert_table_exists(file_path: Path, table_name: str):
    """Assert that a table exists in the workbook."""
    helper = ExcelHelper()
    assert helper.table_exists(file_path, table_name), \
        f"Table '{table_name}' not found in workbook"


def assert_named_range_exists(file_path: Path, range_name: str):
    """Assert that a named range exists."""
    helper = ExcelHelper()
    assert helper.named_range_exists(file_path, range_name), \
        f"Named range '{range_name}' not found in workbook"

